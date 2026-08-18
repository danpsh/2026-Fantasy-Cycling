#!/usr/bin/env python3
"""
Backfill Vuelta a España stage results (top-12 per stage) into a combined,
Year-tagged workbook — vuelta-results-2023-2025.xlsx — matching the shape of
tdf-results-2023-2025.xlsx (Year, Date, Stage, 1st..10th, GC #1-10,
Points/Mountain/Youth #1-3).

Run once per past year (race is over, so all 21 stages exist):
    YEAR=2023 python scripts/scrape_vuelta_history.py
    YEAR=2024 python scripts/scrape_vuelta_history.py
    YEAR=2025 python scripts/scrape_vuelta_history.py

Each run appends/replaces that year's rows in OUT, leaving other years intact.
Names are NOT canonicalized to a roster (there's no historical startlist) — PCS's
own display name is kept as-is, title-cased.

GC/jersey columns are left BLANK, same reason as the live scraper: PCS renders
those classification tabs with JavaScript, so a static-HTML scrape only reliably
gets the stage podium. If you need historical GC/jerseys, fill them by hand from
PCS's site (los-t.php pages show final GC easily since the race is over).
"""
import os
import re
import time
import datetime
import unicodedata
import urllib.request
import urllib.parse

from selectolax.parser import HTMLParser
from openpyxl import Workbook, load_workbook

try:
    from curl_cffi import requests as cffi_requests
except Exception:
    cffi_requests = None

YEAR = os.environ.get("YEAR", "").strip()
if not YEAR:
    raise SystemExit("Set YEAR=2023 (or 2024/2025) before running.")
OUT = os.environ.get("OUT", "vuelta-results-2023-2025.xlsx")
RACE = os.environ.get("RACE_SLUG", "vuelta-a-espana")
MAX_STAGES = int(os.environ.get("MAX_STAGES", "21"))
SCRAPER_API_KEY = os.environ.get("SCRAPER_API_KEY", "").strip()
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"

MONTHS = {m: i for i, m in enumerate(
    ["january", "february", "march", "april", "may", "june", "july",
     "august", "september", "october", "november", "december"], start=1)}

print(f"=== ENV === YEAR={YEAR} OUT={OUT} RACE={RACE}")
print("=== FETCH === via ScraperAPI proxy" if SCRAPER_API_KEY else "=== FETCH === direct (no proxy key)")


def pkey(s):
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9]", "", s.lower())


def name_from_href(href, anchor_text):
    """PCS sometimes glues SURNAME (all caps) directly onto Given Name with no
    separating space/newline in the extracted text (e.g. "CORTINAIván" for
    "García Cortina" + "Iván"). Detect the surname/given-name boundary via case:
    a run of uppercase letters immediately followed by [Upper][lower...] marks
    where a given name starts, even mid-token."""
    t = (anchor_text or "").strip()
    if not t:
        m = re.search(r"rider/([^/?#\"]+)", href or "")
        if not m:
            return ""
        t = m.group(1).replace("-", " ")
    surnames, givens = [], []
    for tok in t.split():
        lm = re.search(r"[a-zà-ÿ]", tok)
        if lm and lm.start() > 0:
            cut = lm.start() - 1
            surn, giv = tok[:cut], tok[cut:]
            if surn:
                surnames.append(surn.capitalize())
            givens.append(giv[:1].upper() + giv[1:])
        elif lm and lm.start() == 0:
            givens.append(tok[:1].upper() + tok[1:])
        else:
            surnames.append(tok.capitalize())
    given_full = " ".join(givens)
    surname_full = " ".join(surnames)
    return (given_full + " " + surname_full).strip() if given_full else surname_full


def fetch(url, tries=3):
    full = "https://www.procyclingstats.com/" + url
    headers = {"User-Agent": UA, "Referer": "https://www.procyclingstats.com/"}
    if SCRAPER_API_KEY:
        target = "https://api.scraperapi.com/?" + urllib.parse.urlencode(
            {"api_key": SCRAPER_API_KEY, "url": full, "country_code": "us"})
        req_headers = {"User-Agent": UA}
    else:
        target = full
        req_headers = headers
    last = None
    for i in range(tries):
        try:
            if cffi_requests is not None:
                r = cffi_requests.get(target, headers=req_headers, impersonate="chrome", timeout=90)
                if r.status_code == 200:
                    return r.text
                last = f"HTTP {r.status_code}"
                if r.status_code in (401, 403) and SCRAPER_API_KEY:
                    raise RuntimeError(f"proxy rejected ({r.status_code}) — likely out of API credits; not retrying")
            else:
                req = urllib.request.Request(target, headers=req_headers)
                with urllib.request.urlopen(req, timeout=90) as resp:
                    return resp.read().decode("utf-8", "replace")
        except RuntimeError:
            raise
        except Exception as e:
            last = str(e)
        time.sleep(5 * (i + 1))
    raise RuntimeError(f"fetch failed for {url}: {last}")


def table_names(table):
    body = table.css_first("tbody") or table
    out, seen = [], set()
    for tr in body.css("tr"):
        a = tr.css_first('a[href*="rider/"]')
        if not a:
            continue
        nm = name_from_href(a.attributes.get("href", ""), a.text(strip=True))
        if nm and nm not in seen:
            seen.add(nm)
            out.append(nm)
    return out


def stage_table(html):
    """Largest results table on the base stage page = that day's finish order."""
    best = []
    for t in HTMLParser(html).css("table.results"):
        names = table_names(t)
        if len(names) > len(best):
            best = names
    return best


def parse_date(html, n):
    m = re.search(r"(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})", html)
    if m and m.group(2).lower() in MONTHS:
        return f"{int(m.group(3)):04d}-{MONTHS[m.group(2).lower()]:02d}-{int(m.group(1)):02d}"
    return ""


def pad(names, n):
    return (names[:n] + [""] * n)[:n]


def classify_lists(html):
    """Largest results table(s) on a PCS page, in DOM order (kept if >=10 rows)."""
    kept = []
    for t in HTMLParser(html).css("table.results"):
        names = table_names(t)
        if len(names) >= 10:
            kept.append(names)
    return kept


def fetch_final_standings():
    """Final GC + jersey standings, fetched from PCS's dedicated classification
    pages (race/.../gc, /points, /kom, /youth) — these render statically even
    though the per-stage tabs need JS, since they ARE the whole page's content."""
    result = {"gc": [], "points": [], "kom": [], "youth": []}
    for key, suffix in [("gc", "gc"), ("points", "points"), ("kom", "kom"), ("youth", "youth")]:
        try:
            html = fetch(f"race/{RACE}/{YEAR}/{suffix}")
        except Exception as e:
            print(f"final {key}: fetch error {e}")
            continue
        lists = classify_lists(html)
        if lists:
            result[key] = lists[0]
            print(f"final {key}: {len(lists[0])} riders, #1 {lists[0][0]}")
        else:
            print(f"final {key}: no table found")
    return result


NGC, NJER = 10, 3
NPLACE = 10
HEADER = (["Year", "Date", "Stage"]
          + ["1st", "2nd", "3rd", "4th", "5th", "6th", "7th", "8th", "9th", "10th"]
          + [f"GC #{i}" for i in range(1, NGC + 1)]
          + [f"Points #{i}" for i in range(1, NJER + 1)]
          + [f"Mountain #{i}" for i in range(1, NJER + 1)]
          + [f"Youth #{i}" for i in range(1, NJER + 1)])


def load_existing_other_years():
    """Read every row from a prior year (kept as-is) so this run only touches YEAR."""
    rows = []
    if os.path.exists(OUT):
        try:
            wb = load_workbook(OUT)
            ws = wb.active
            for r in list(ws.iter_rows(values_only=True))[1:]:
                if r and str(r[0]) != YEAR:
                    rows.append(list(r))
            print(f"kept {len(rows)} row(s) from other years in {OUT}")
        except Exception as e:
            print(f"could not read existing {OUT}: {e}")
    return rows


def main():
    other_rows = load_existing_other_years()
    year_rows = []
    for n in range(1, MAX_STAGES + 1):
        try:
            html = fetch(f"race/{RACE}/{YEAR}/stage-{n}")
        except Exception as e:
            print(f"stage {n}: fetch error {e}")
            continue
        names = stage_table(html)
        if not names:
            print(f"stage {n}: no table (stage {n} may not exist this year — stopping)")
            break
        row = [YEAR, parse_date(html, n), n] + pad(names, NPLACE) + [""] * (NGC + 3 * NJER)
        year_rows.append(row)
        print(f"stage {n}: {row[1]} | win {row[3]}")

    if not year_rows:
        print("\nNothing scraped; leaving existing file untouched.")
        return

    # Final GC + jersey standings (from the race-is-over classification pages) go
    # on the LAST scraped stage's row only — that is the snapshot the scoring
    # engine reads (GC/jerseys count once, at the latest stage).
    final = fetch_final_standings()
    last = year_rows[-1]
    last[3 + NPLACE:3 + NPLACE + NGC] = pad(final["gc"], NGC)
    last[3 + NPLACE + NGC:3 + NPLACE + NGC + NJER] = pad(final["points"], NJER)
    last[3 + NPLACE + NGC + NJER:3 + NPLACE + NGC + 2 * NJER] = pad(final["kom"], NJER)
    last[3 + NPLACE + NGC + 2 * NJER:3 + NPLACE + NGC + 3 * NJER] = pad(final["youth"], NJER)

    all_rows = other_rows + year_rows
    all_rows.sort(key=lambda r: (str(r[0]), int(r[2]) if str(r[2]).isdigit() else 0))
    wb = Workbook(); ws = wb.active; ws.title = "Results"; ws.append(HEADER)
    for r in all_rows:
        ws.append(r)
    wb.save(OUT)
    print(f"\nWrote {len(all_rows)} total row(s) to {OUT} ({len(year_rows)} for {YEAR} this run)")


if __name__ == "__main__":
    main()
