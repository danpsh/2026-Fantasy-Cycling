#!/usr/bin/env python3
"""
Scrape Tour de France results from ProCyclingStats -> tdf-results.xlsx.

Parsed directly with selectolax (procyclingstats 0.2.8 can't read current PCS).
Each classification has its own page; on each we take the LARGEST results table
(the full standings, ignoring small "today" widgets). Riders are identified by
their URL slug and mapped to canonical roster spelling from tdf-startlist.js.

    pip install procyclingstats openpyxl   # (brings selectolax)
    YEAR=2025 OUT=tdf-results-2025-test.xlsx python scripts/scrape_tdf.py
"""
import os
import re
import json
import time
import datetime
import unicodedata
import urllib.request
import urllib.parse

from selectolax.parser import HTMLParser
from openpyxl import Workbook, load_workbook

# curl_cffi impersonates a real Chrome TLS/HTTP2 fingerprint, which gets past
# ProCyclingStats' Cloudflare block (plain urllib gets 403 from datacenter IPs).
try:
    from curl_cffi import requests as cffi_requests
except Exception:
    cffi_requests = None

YEAR = os.environ.get("YEAR", "2026")
OUT = os.environ.get("OUT", "tdf-results.xlsx")
RACE = os.environ.get("RACE_SLUG", "tour-de-france")
MAX_STAGES = int(os.environ.get("MAX_STAGES", "21"))
UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"

# ProCyclingStats sits behind Cloudflare, which 403s datacenter IPs (i.e. GitHub
# Actions runners). When SCRAPER_API_KEY is set we route each request through
# ScraperAPI's residential IP pool to get a clean IP; when it's absent we hit PCS
# directly (works from a home/residential IP for local runs). Free tier is ~1,000
# requests/month; a full 21-stage scrape is ~21 requests.
SCRAPER_API_KEY = os.environ.get("SCRAPER_API_KEY", "").strip()

SCHED_2026 = {1: "2026-07-04", 2: "2026-07-05", 3: "2026-07-06", 4: "2026-07-07",
              5: "2026-07-08", 6: "2026-07-09", 7: "2026-07-10", 8: "2026-07-11",
              9: "2026-07-12", 10: "2026-07-14", 11: "2026-07-15", 12: "2026-07-16",
              13: "2026-07-17", 14: "2026-07-18", 15: "2026-07-19", 16: "2026-07-21",
              17: "2026-07-22", 18: "2026-07-23", 19: "2026-07-24", 20: "2026-07-25",
              21: "2026-07-26"}
MONTHS = {m: i for i, m in enumerate(
    ["january", "february", "march", "april", "may", "june", "july",
     "august", "september", "october", "november", "december"], start=1)}

print(f"=== ENV === YEAR={YEAR} OUT={OUT} RACE={RACE}")
print("=== FETCH === via ScraperAPI proxy" if SCRAPER_API_KEY else "=== FETCH === direct (no proxy key)")


def pkey(s):
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9]", "", s.lower())


def load_canon():
    # tdf-startlist.js is a JS object literal (unquoted keys), not JSON, so pull
    # rider names with a regex rather than json.loads.
    try:
        with open("tdf-startlist.js", "r", encoding="utf-8") as f:
            t = f.read()
        names = re.findall(r'(?:"name"|name)\s*:\s*"([^"]+)"', t)
        canon = {pkey(n): n for n in names}
        if not canon:
            raise ValueError("no rider names found")
        print(f"loaded {len(canon)} canonical names from tdf-startlist.js")
        return canon
    except Exception as e:
        print("tdf-startlist.js NOT loaded (names will be slug-cased):", e)
        return {}


CANON = load_canon()
# Sorted longest-first for the prefix fallback below.
CANON_ITEMS = sorted(CANON.items(), key=lambda kv: -len(kv[0]))


def canon_lookup(slug_key):
    """Map a PCS name-key to the roster's canonical spelling.

    PCS frequently carries extra surname parts the roster omits
    ("juan-ayuso-pesquera" vs roster "Juan Ayuso"). After an exact miss, accept a
    canonical name that is a prefix of the slug key (or vice versa) — but only if
    EXACTLY ONE canonical matches, so we never guess between two similar riders.
    """
    if slug_key in CANON:
        return CANON[slug_key]
    uniq = []
    for k, nm in CANON_ITEMS:
        if len(k) >= 6 and (slug_key.startswith(k) or k.startswith(slug_key)):
            if nm not in uniq:
                uniq.append(nm)
    return uniq[0] if len(uniq) == 1 else None


def name_from_href(href):
    m = re.search(r"rider/([^/?#\"]+)", href or "")
    if not m:
        return ""
    slug = m.group(1).replace("-", " ").strip()
    return canon_lookup(pkey(slug)) or " ".join(w.capitalize() for w in slug.split())


def fetch(url, tries=5):
    full = "https://www.procyclingstats.com/" + url
    headers = {
        "User-Agent": UA,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Referer": "https://www.procyclingstats.com/",
    }
    # When a proxy key is present, hand PCS's URL to ScraperAPI and fetch THAT
    # instead. ScraperAPI supplies a residential IP + solves Cloudflare, then
    # returns the target page's HTML. render=false (default) is enough for PCS's
    # server-rendered tables; country_code=us keeps the IP pool consistent.
    if SCRAPER_API_KEY:
        target = "https://api.scraperapi.com/?" + urllib.parse.urlencode(
            {"api_key": SCRAPER_API_KEY, "url": full, "country_code": "us"})
        req_headers = {"User-Agent": UA}
    else:
        target = full
        req_headers = headers

    last = None
    for i in range(tries):
        try:
            if cffi_requests is not None:
                r = cffi_requests.get(target, headers=req_headers, impersonate="chrome", timeout=70)
                if r.status_code == 200:
                    return r.text
                last = f"HTTP {r.status_code}"
            else:
                req = urllib.request.Request(target, headers=req_headers)
                with urllib.request.urlopen(req, timeout=70) as resp:
                    return resp.read().decode("utf-8", "replace")
        except Exception as e:
            last = str(e)
        # ScraperAPI 5xx / timeouts are usually transient (PCS load right after a
        # stage, momentary proxy failure) — back off progressively and retry.
        time.sleep(6 * (i + 1))
    raise RuntimeError(f"fetch failed for {url}: {last}")


# ---- accented display names for the annual-league file ----
# annual-results.xlsx shows rider names verbatim, so we keep their accents
# ("Tadej Poga\u010dar"). App-side scoring is accent- and case-insensitive but word
# ORDER matters, so every stored display name must be First-Last AND reduce to the
# same letter key as the roster's canonical spelling.
DISPLAY = {}   # pkey(name) -> accented "First Last" display


def register_display(raw_text, canonical):
    """Record an accented spelling for a canonical name, but only when it is
    provably safe: accept the PCS anchor text (or a word-reordering of it) only if
    it reduces to the SAME letter key as the canonical name. That guarantees the
    stored name still matches the roster after the app normalizes it; otherwise we
    keep nothing and callers fall back to the plain canonical spelling."""
    if not raw_text or not canonical:
        return
    ck = pkey(canonical)
    toks = raw_text.split()
    cands = [raw_text]
    if len(toks) >= 2:
        cands.append(" ".join([toks[-1]] + toks[:-1]))   # PCS "Last First" -> "First Last"
        cands.append(" ".join(toks[1:] + [toks[0]]))      # the reverse, just in case
    for c in cands:
        if pkey(c) == ck:
            DISPLAY.setdefault(ck, c)
            return


def display_name(canonical):
    return DISPLAY.get(pkey(canonical), canonical)


def table_names(table):
    body = table.css_first("tbody") or table
    out, seen = [], set()
    for tr in body.css("tr"):
        a = tr.css_first('a[href*="rider/"]')
        if not a:
            continue
        nm = name_from_href(a.attributes.get("href", ""))
        if nm and nm not in seen:
            seen.add(nm)
            out.append(nm)
            register_display(a.text(strip=True), nm)
    return out


def classify(html, debug=False):
    """Return {stage,gc,points,kom,youth} name-lists.

    PCS puts all five standings on one stage page, in DOM order
    stage, GC, points, KOM, youth, interleaved with small daily sub-tables.
    The five cumulative standings are the first five results tables with
    enough riders (>=20); the daily sub-results (n<20) are skipped.
    """
    kept = []
    for i, t in enumerate(HTMLParser(html).css("table.results")):
        names = table_names(t)
        if debug:
            print(f"    [{i}] n={len(names)} top3={names[:3]}")
        if len(names) >= 20:
            kept.append(names)
    order = ["stage", "gc", "points", "kom", "youth"]
    return {k: (kept[i] if i < len(kept) else []) for i, k in enumerate(order)}


def parse_date(html, n):
    m = re.search(r"(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})", html)
    if m and m.group(2).lower() in MONTHS:
        return f"{int(m.group(3)):04d}-{MONTHS[m.group(2).lower()]:02d}-{int(m.group(1)):02d}"
    return SCHED_2026.get(n, "") if YEAR == "2026" else ""


def pad(names, n):
    return (names[:n] + [""] * n)[:n]


HEADER = (["Date", "Stage"]
          + ["1st", "2nd", "3rd", "4th", "5th", "6th",
             "7th", "8th", "9th", "10th", "11th", "12th"])
NPLACE = 12  # TdF scores the top 12 stage finishers

# ---- annual-league mirror ----
# The Tour is one race inside the season-long annual league, whose results live in
# annual-results.xlsx with a different shape: Excel date-serial | "Tour de France"
# | "Stage N" | top 10. Stage 1 (the TTT) is deliberately NOT scored in the annual
# competition, so it is skipped. Only 'Tour de France' rows are ever touched.
ANNUAL_OUT = os.environ.get("ANNUAL_OUT", "annual-results.xlsx")
ANNUAL_RACE = "Tour de France"
ANN_NPLACE = 10
EXCEL_EPOCH = datetime.date(1899, 12, 30)


def to_serial(datestr):
    """ISO 'YYYY-MM-DD' -> Excel date serial (int); pass anything else through."""
    try:
        y, m, d = (int(x) for x in str(datestr).split("-"))
        return (datetime.date(y, m, d) - EXCEL_EPOCH).days
    except Exception:
        return datestr


def norm_row(r):
    """Coerce any stored row to the [Date, Stage, 1st..12th] shape."""
    r = list(r)
    date = r[0] if len(r) > 0 else ""
    stage = r[1] if len(r) > 1 else ""
    places = [("" if c is None else c) for c in r[2:2 + NPLACE]]
    return [date, stage] + pad(places, NPLACE)


def load_existing():
    """Read prior stages from OUT so a single-stage run doesn't wipe them."""
    out = {}
    if os.path.exists(OUT):
        try:
            wb = load_workbook(OUT)
            ws = wb.active
            for row in list(ws.iter_rows(values_only=True))[1:]:
                if row and row[1] not in (None, ""):
                    try:
                        out[int(row[1])] = norm_row(row)
                    except (ValueError, TypeError):
                        pass
            print(f"loaded {len(out)} existing stage(s) from {OUT}")
        except Exception as e:
            print(f"could not read existing {OUT}: {e}")
    return out


def stages_to_scrape():
    """Which stage(s) to fetch this run.

    Default: only the stage scheduled for today's date (UTC), so daily runs do
    one request instead of re-scraping all 21. Overrides via STAGE env:
      STAGE=8     -> just stage 8 (manual backfill of a single stage)
      STAGE=1,2   -> stages 1 and 2 (comma list, for testing)
      STAGE=all   -> every stage 1..MAX_STAGES (full rebuild / testing old years)
    """
    forced = os.environ.get("STAGE", "").strip().lower()
    if forced == "all":
        return list(range(1, MAX_STAGES + 1))
    if forced:
        return [int(x) for x in re.split(r"[,\s]+", forced) if x]
    today = datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%d")
    rev = {v: k for k, v in SCHED_2026.items()}
    n = rev.get(today)
    return [n] if n else []


def update_annual(stages):
    """Mirror the TdF stages into annual-results.xlsx (skipping Stage 1, the TTT).

    Every non-Tour row the user maintains by hand is preserved exactly — only
    'Tour de France' rows are rewritten, from the full merged stage set. Names
    carry accents when a safe accented spelling is known (seeded from the existing
    Tour rows so hand-entered accents survive, plus anything captured live)."""
    if not os.path.exists(ANNUAL_OUT):
        print(f"annual: {ANNUAL_OUT} not found; skipping.")
        return
    try:
        wb = load_workbook(ANNUAL_OUT)
    except Exception as e:
        print(f"annual: could not open {ANNUAL_OUT}: {e}; skipping.")
        return
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("annual: empty sheet; skipping.")
        return
    header = rows[0]

    def col(name):
        for i, h in enumerate(header):
            if h is not None and str(h).strip().lower() == name.lower():
                return i
        return None

    ci_race = col("Race Name")
    ci_names = [col(c) for c in ["1st", "2nd", "3rd", "4th", "5th", "6th", "7th", "8th", "9th", "10th"]]
    if ci_race is None:
        print("annual: no 'Race Name' column; skipping.")
        return

    def is_tdf(r):
        return r[ci_race] is not None and str(r[ci_race]).strip() == ANNUAL_RACE

    # Seed accented spellings from the existing Tour rows (so Stage 2/3 accents
    # entered by hand survive even though only today's stage is re-scraped live).
    for r in rows[1:]:
        if is_tdf(r):
            for ci in ci_names:
                if ci is not None and ci < len(r) and r[ci]:
                    nm = str(r[ci]).strip()
                    DISPLAY.setdefault(pkey(nm), nm)

    # Delete existing Tour rows bottom-up (keeps every other row + its formatting).
    idx = [i for i, r in enumerate(ws.iter_rows(values_only=True), start=1)
           if i > 1 and is_tdf(r)]
    for i in reversed(idx):
        ws.delete_rows(i, 1)

    # Append fresh Tour rows from the merged stage set (skip Stage 1 / TTT).
    added = 0
    for n in sorted(stages):
        if n < 2:
            continue
        row = stages[n]
        date_iso, names = row[0], row[2:2 + NPLACE]
        placed = [display_name(x) for x in names[:ANN_NPLACE]]
        placed = (placed + [""] * ANN_NPLACE)[:ANN_NPLACE]
        ws.append([to_serial(date_iso), ANNUAL_RACE, f"Stage {n}"] + placed)
        added += 1
    wb.save(ANNUAL_OUT)
    print(f"annual: mirrored {added} Tour stage row(s) into {ANNUAL_OUT} (Stage 1 skipped)")


def main():
    targets = stages_to_scrape()
    if not targets:
        today = datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%d")
        print(f"\nNo stage scheduled for {today}; nothing to do.")
        return
    print(f"stages to scrape: {targets}")

    stages = load_existing()
    # Idempotency: in default (today) mode, if the target stage is already
    # recorded, do nothing — so running every 15 min in the finish window is
    # cheap. The first run that catches results commits; later runs exit here
    # with no proxy call and no commit. A forced STAGE always re-scrapes.
    forced = bool(os.environ.get("STAGE", "").strip())
    if not forced:
        pending = [n for n in targets if n not in stages]
        if not pending:
            print(f"stage {targets} already recorded; nothing to do.")
            return
        targets = pending
    scraped = 0
    for n in targets:
        # Stage 1 is a TTT: on the base stage page the main table is the TEAM
        # result, and the individual-GC table there is unreliable to locate by
        # position. Use PCS's dedicated per-stage GC page instead — its main table
        # IS the individual GC order (league rule: TTT placements = GC order).
        if n == 1:
            b = f"race/{RACE}/{YEAR}/stage-{n}-gc"
        else:
            b = f"race/{RACE}/{YEAR}/stage-{n}"
        try:
            html = fetch(b)
        except Exception as e:
            print(f"stage {n}: fetch error {e}")
            continue
        names = classify(html)["stage"]  # kept[0] = the page's main standings
        if not names:
            print(f"stage {n}: no results table yet (stage not finished?)")
            continue
        if n == 1:
            print("stage 1 (TTT): using dedicated GC page for individual placements")
        row = [parse_date(html, n), n] + pad(names, NPLACE)
        stages[n] = row
        scraped += 1
        print(f"stage {n}: {row[0]} | win {row[2]} | 12th {row[13]}")

    if scraped == 0:
        print("\nNothing scraped; leaving existing file untouched.")
        return

    wb = Workbook(); ws = wb.active; ws.title = "Results"; ws.append(HEADER)
    for k in sorted(stages):
        ws.append(stages[k])
    wb.save(OUT)
    print(f"\nWrote {len(stages)} stage(s) to {OUT} ({scraped} new/updated this run)")

    # Mirror into the annual league — only for the live default file/year, so
    # year-2025 test runs (custom OUT) never disturb annual-results.xlsx.
    if OUT == "tdf-results.xlsx" and YEAR == "2026":
        update_annual(stages)


if __name__ == "__main__":
    main()
